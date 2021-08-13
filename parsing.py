import os
import zipfile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


class Parser:
    def __init__(self, start_directory, result_directory, text_to_replace, result_text, message_screen=None):
        self.start_directory = start_directory
        self.result_directory = result_directory
        self.text_to_replace = text_to_replace
        self.result_text = result_text
        self.message_screen = message_screen
        self.is_replaced = False                        # Были ли изменения в файле

    @staticmethod
    def show_directory(path):
        """ Генератор пробегается по файлам в директории
        :param path: Путь к директории
        """
        files = os.listdir(path)
        for file in files:
            yield file

    def parse_excel(self, start_path, result_path):
        """ Парсинг файла excel
        :param start_path: Папка которую парсим
        :param result_path: Папка куда сохраняем
        :return 1 - успех, 0 - ошибка
        """
        self.is_replaced = False
        try:
            workbook = load_workbook(start_path)
            for sheet in workbook.sheetnames:
                worksheet = workbook[sheet]
                self.parse_sheet(worksheet)
                if self.is_replaced:
                    workbook.save(result_path)
                    return 1
        except InvalidFileException:
            return 0
        except zipfile.BadZipFile:
            return 0

    def parse_sheet(self, worksheet):
        """ Парсинг книги из файла
        :param worksheet: Книга из файла excel
        :return:
        """
        for row in worksheet.iter_rows():
            for cell in row:
                cell_text = str(cell.value)
                if self.text_to_replace in cell_text:
                    new_text = cell_text.replace(self.text_to_replace, self.result_text)
                    cell.value = new_text
                    self.is_replaced = True

    def run(self):
        """ Старт парсера """
        try:
            for file_name in self.show_directory(self.start_directory):
                if self.message_screen:
                    self.message_screen.appendHtml("Чтение файла <b>{}</b>".format(file_name))
                else:
                    print("Чтение файла <b>{}</b>".format(file_name))
                path = os.path.join(self.start_directory, file_name)
                result_path = os.path.join(self.result_directory, file_name)
                parsing_result = self.parse_excel(path, result_path)
                self.show_parsing_results(parsing_result)
            return 'Парсинг завершен успешно!'
        except FileNotFoundError:
            return 'Ошибка парсинга. Указанная папка не обнаружена!'

    def show_parsing_results(self, parsing_result):
        """ Отображение результатов парсинга после каждого обработанного файла
        :param parsing_result: - код результата
        """
        if parsing_result:
            if self.message_screen:
                self.message_screen.appendHtml('<span style="color: green;">Заменено</span>')
            else:
                print('REPLACED')
        else:
            if self.message_screen:
                self.message_screen.appendHtml(
                    '<span style="color: red;">Не могу заменить из-за старого формата Excel</span>')
            else:
                print('OLD FORMAT')


if __name__ == '__main__':
    parser = Parser(
        start_directory='documents',
        result_directory='new_documents',
        text_to_replace='to replace',
        result_text='replaced')
    result = parser.run()
    print(result)
